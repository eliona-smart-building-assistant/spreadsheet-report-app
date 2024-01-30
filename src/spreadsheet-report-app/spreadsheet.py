import os
import json
from json import JSONDecoder
from openpyxl import load_workbook
import formulas
import pandas as pd
import shutil
import utils.logger as log
from datetime import datetime, timedelta
import pytz

from eliona_modules.api.core.eliona_core import ElionaApiHandler, ConStat

LOGGER_NAME = "Spreadsheet"
LOGGER_LEVEL = log.LOG_LEVEL_DEBUG
FILL_UP = True #If true will first fill up with heading and then tailing none lines

class Spreadsheet:

	logger = log.createLogger(LOGGER_NAME, loglevel=LOGGER_LEVEL)

	def __init__(self, logLevel:int=log.LOG_LEVEL_DEBUG) -> None:
		"""
		Initialize the class
		"""

		self.reportFilePath = ""
		self.logger.setLevel(logLevel)

	def createReport(self, startDt:datetime, endDt:datetime, connectionSettings:dict, reportSettings:dict) -> bool:
		"""
		Create the requested report

		Params
		------ 
		startDt:datetime			= Start time of the Report
		endDt:datetime 				= End Time of teh Report => Will include the given end time by the data point as last entry
		connectionSettings:dict 	= Connection settings for the eliona handler {"host", "api", "projectId", "apiKey", "dbTimeZone"}
		reportSettings:dict			= Settings for the to configured report

		Return
		----
		bool -> Will return True if report was successfully created. // False if a error occurred

		"""

		self.reportFilePath = reportSettings["tempPath" ]
		
		# If the file already exists we skip the creation part
		if os.path.isfile(self.reportFilePath):
			return True

		#set the local variables
		_reportCreatedSuccessfully = False

		self.logger.debug("--------connect--------")
		self.logger.debug("Host: " + str(connectionSettings["host"]))

		#Connect to the eliona instance
		eliona = ElionaApiHandler(settings=connectionSettings, logger=LOGGER_NAME)
		eliona.check_connection() 

		#Check if the connection is established
		if eliona.connection == ConStat.CONNECTED:
				
			self.logger.debug("--------Create Table--------")
			#Call the report creator
			if (reportSettings["type"] == "DataListSequential") or (reportSettings["type"] == "DataListParallel"):
				_reportCreatedSuccessfully = self.__createDataListReport( eliona=eliona, settings=reportSettings, startDateTime=startDt, endDateTime=endDt)
			elif reportSettings["type"] == "DataEntry":
				_reportCreatedSuccessfully = self.__createDataEntryReport(eliona=eliona, settings=reportSettings, startDateTime=startDt, endDateTime=endDt)
			else:
				self.logger.error(f"Invalid report configuration: reports['type']={reportSettings['type']}")
		else:
			self.logger.info("Connection not possible. Will try again.")

		return _reportCreatedSuccessfully

	def __createDataEntryReport(self, eliona:ElionaApiHandler, settings:dict, startDateTime:datetime, endDateTime:datetime) -> bool:
		"""
		Create the table report from the given template

		settings:dict = Settings dictionary 
		"""
		_reportCreated = False

		#Only copy the template if needed 
		if settings["fromTemplate"]:
			shutil.copyfile(src=settings["templateFile"], dst=self.reportFilePath)

		#Read the template 
		_dataTable = self.__readTableTemplate(settings=settings)

		for _rowIndex, _row in _dataTable.iterrows(): #iterate over rows

			_timeStampFormat = ""

			for _columnIndex, _value in _row.items():
				_jsonFound = False
				_newValue = _value
				#Search for json config
				#{"assetId":"xxx", "attribute":"yyy"}
				if type(_value) == str:

					for _config, _configRaw in self.__findJson(_value):
						_jsonFound = True

						if ("timeStampStart" in _config):
						
							_timeStampFormat = _config["timeStampStart"]
							_newValue = startDateTime.strftime(_timeStampFormat)

						elif ("timeStampEnd" in _config):

							_timeStampFormat = _config["timeStampEnd"]
							_newValue = (endDateTime- timedelta(days=1)).strftime(_timeStampFormat)

						elif ((("assetId" in _config) or ("assetGai" in _config)) and ("attribute" in _config)):

							_timeStampFormat = "%Y-%m-%d %H:%M:%S"

							if "assetId" in _config: 
								_assetId = int(_config["assetId"])
							else:
								_assetId = 0

							if "assetGai" in _config:
								_assetGai = _config["assetGai"]
							else:
								_assetGai = ""

							_data, _dataFrame, _correctTimestamps = self.__getAggregatedDataList(	eliona=eliona, 
																									assetGai=_assetGai,
																									assetId=_assetId, 
																									attribute=str(_config["attribute"]), 
																									startDateTime= startDateTime, 
																									endDateTime=startDateTime + timedelta(days=1),
																									raster=_config["raster"],
																									mode=_config["mode"],
																									timeStampKey="TimeStamp",
																									valueKey="Value")


							#Set the TimeStamp straight
							_dataFrame["TimeStamp"] = pd.to_datetime(arg=_dataFrame["TimeStamp"]).dt.strftime(_timeStampFormat)						

							#Just get the right timestamp
							_dataFrame = _dataFrame[(_dataFrame["TimeStamp"] == startDateTime.strftime(_timeStampFormat))]

							#Set the Value to the Spreadsheet cell
							if(len(_dataFrame.index) == 1):
								_newValue = _newValue.replace(_configRaw,  str(_dataFrame.at[0,"Value"]))
							elif(len(_dataFrame.index) > 1):
								self.logger.error("Received more than one data entry from the database: " + str(len(_dataFrame.index)))
								_newValue = _newValue.replace(_configRaw, "DOUBLE-VALUE")
							else:
								
								# Get the config what to enter if no value was found
								_noValue = _config.get("fillNone", "NO-VALUE")

								if _noValue == "last":
									_filler = self.getLastReceivedValue(eliona=eliona, assetGai=_assetGai, assetId=_assetId, attribute=str(_config["attribute"]), startDateTime=startDateTime, timeStampFormat=_timeStampFormat)
									dateTimeStr = startDateTime.isoformat()
									assetAttribute = str(_config["attribute"])
									self.logger.warning(f"No value found and was replaced by last value from year raster. Asset: {_assetGai}, Attribute: {assetAttribute}, DateTime: {dateTimeStr}")
								elif _noValue == "zero":
									_filler = 0
								else:
									_filler = "NO-VALUE"

								# Replace the json config with the filler value
								_newValue = _newValue.replace(_configRaw, str(_filler))

				if _jsonFound:
					if self.__isFloat(_newValue):
						_dataTable.at[_rowIndex, _columnIndex] = float(_newValue)
					else:
						_dataTable.at[_rowIndex, _columnIndex] = _newValue



		#Write the Data to the 
		_reportCreated = self.__writeDataToFile(data=_dataTable, settings=settings)

		return _reportCreated

	def __createDataListReport(self, eliona:ElionaApiHandler, settings:dict, startDateTime:datetime, endDateTime:datetime) -> bool:
		"""
		Create the table report from the given template

		eliona:ElionaApiHandler = Eliona handler
		settings:dict = Settings dictionary 
		"""

		#Init the data
		_timeStampColumnName = ""
		_timeStampFormat = ""
		_reportCreated = False
		_correctTimestamps = False

		#Only copy the template if needed 
		if settings["fromTemplate"]:
			shutil.copyfile(src=settings["templateFile"], dst=self.reportFilePath)

		#Read the template 
		_dataTable = self.__readTableTemplate(settings=settings)

		#Read the configuration from the template
		_configDict = {}
		for _columnName in _dataTable.columns.values:

			_firstRow = int(settings["firstRow"])
			_cellData = _dataTable.at[_firstRow, _columnName]

			if type(_dataTable.at[_firstRow, _columnName]) == str:
				_config = json.loads(_cellData)
				_configDict[_columnName] = _config

		
		dataColumns = []
		#Search for the time stamp configuration
		for _columnName in _configDict:

			if "timeStamp" in _configDict[_columnName]:
				
				_timeStampColumnName = _columnName #Save the timestamp key
				_timeStampFormat = _configDict[_columnName]["timeStamp"]

				_raster = str(_configDict[_columnName]["raster"])

				#Get the time tick 
				if _raster == "MONTH":

					#If we got monthly spans we need to work with replace instead of timedelta.
					#And we will always start with 
					startDateTime = startDateTime.replace(month=1)

				elif _raster.startswith("H"):

					timeTick = timedelta(hours=int(_raster.removeprefix("H")))	

				elif _raster.startswith("M"):

					timeTick = timedelta(minutes=int(_raster.removeprefix("M")))	

				elif _raster.startswith("S"):

					timeTick = timedelta(seconds=int(_raster.removeprefix("S")))

				else:
					self.logger.error("No valid time span found")
					return False
				
				#Create the time column with the required timestamp
				_timeStampList = []
				_timeStamp = startDateTime

				# Define the end of the list. We do not want to fiddle around with the time interval. Therefore just -1 minute :-) 
				_endTimeStampForList = (endDateTime-timedelta(minutes=1))

				if _raster == "MONTH":
					while (_timeStamp <= _endTimeStampForList):
						_timeStampList.append((_timeStamp).strftime(_timeStampFormat))

						if _timeStamp.month + 1 > 12:
							_timeStamp = _timeStamp.replace(month = 1)
							_timeStamp = _timeStamp.replace(year= _timeStamp.year + 1)
						else:
							_timeStamp = _timeStamp.replace(month=_timeStamp.month + 1)

				else:
					while (_timeStamp <= _endTimeStampForList):
						_timeStampList.append((_timeStamp).strftime(_timeStampFormat))
						_timeStamp = _timeStamp + timeTick

				#Create the DataFrame for the readed Data
				_dataTable = pd.DataFrame( _timeStampList, columns=[_timeStampColumnName])

			elif ((("assetId" in _configDict[_columnName]) or ("assetGai" in _configDict[_columnName])) 
					and ("attribute" in _configDict[_columnName]) 
					and ("mode" in _configDict[_columnName])):
				
				dataColumns.append(_columnName)

				if "assetId" in _configDict[_columnName]: 
					_assetId = int(_configDict[_columnName]["assetId"])
				else:
					_assetId = 0

				if "assetGai" in _configDict[_columnName]:
					_assetGai = _configDict[_columnName]["assetGai"]
				else:
					_assetGai = ""


				_data, _dataFrame, _correctTimestamps = self.__getAggregatedDataList(	eliona=eliona,
																						assetGai=_assetGai, 
																						assetId=_assetId,
																						attribute=str(_configDict[_columnName]["attribute"]), 
																						startDateTime=startDateTime, 
																						endDateTime=endDateTime,
																						raster=_raster,
																						mode=_configDict[_columnName]["mode"],
																						timeStampKey = _timeStampColumnName,
																						valueKey=_columnName)


				#Convert the data with the right timestamp format
				_dataFrame[_timeStampColumnName] = pd.to_datetime(arg=_dataFrame[_timeStampColumnName]).dt.strftime(_timeStampFormat)
				#Merge the Aggregated data with the current dataframe
				_dataTable = pd.merge(_dataTable, _dataFrame, how='left', on=_timeStampColumnName)

			else:
				self.logger.error("No valid table configuration.")

		# Search for empty cells in the data columns
		#Get the empty data from the table
		emptyTableFrame = pd.isnull(_dataTable)

		for _columnName in dataColumns:

			# Check if empty entries need to be fixed for this column
			if _configDict[_columnName].get("fillNone", ""):

				# Get the Rows with empty cells
				singleDataFrame = emptyTableFrame[emptyTableFrame[_columnName] == True]

				# Check if Data Column has entry cells
				if singleDataFrame.size >= 1:

					#Fix all the empty cells
					for _itemIndex, _columnValues in singleDataFrame.iterrows():

						# Get the configuration of the row from teh template
						fillNone = _configDict[_columnName].get("fillNone", "NO-VALUE")

						if "assetId" in _configDict[_columnName]: 
							_assetId = int(_configDict[_columnName]["assetId"])
						else:
							_assetId = 0

						if "assetGai" in _configDict[_columnName]:
							_assetGai = _configDict[_columnName]["assetGai"]
						else:
							_assetGai = ""
						_assetAttribute = str(_configDict[_columnName]["attribute"])

						# Get the timestamp
						dateTimeStr = _dataTable.at[_itemIndex, _timeStampColumnName]
						_startTimestamp = datetime.strptime(dateTimeStr, _timeStampFormat).astimezone(pytz.timezone("Europe/Zurich"))

						if fillNone == "last":
							_newValue = self.getLastReceivedValue(eliona=eliona, assetGai=_assetGai, assetId=_assetId, attribute=_assetAttribute, startDateTime=_startTimestamp)
							_newValueStr = str(_newValue)
							self.logger.warning(f"No value found and was replaced by last value: {_newValueStr}. AssetGAI: {_assetGai}, Attribute: {_assetAttribute}, DateTime: {dateTimeStr}")
						elif fillNone == "zero":
							_newValue = 0
							self.logger.warning(f"No value found and was replaced by zero. AssetGAI: {_assetGai}, Attribute: {_assetAttribute}, DateTime: {dateTimeStr} ")
						else:
							_newValue = "NO-VALUE" 

						#Set the new Value to the table
						_dataTable.at[_itemIndex, _columnName] = _newValue

		#Fill up the empty cells. First with the newer ones. In case the first row is empty we will also fill with the older ones up
		if "fillNone" in settings:
			
			if settings["fillNone"]:
				_dataTable = _dataTable.fillna(method="ffill")
				_dataTable = _dataTable.fillna(method="bfill")

		else:

			if FILL_UP:
				_dataTable = _dataTable.fillna(method="ffill")
				_dataTable = _dataTable.fillna(method="bfill")

		#Write the data to file
		_reportCreated = self.__writeDataToFile(data=_dataTable, settings=settings)

		return _reportCreated

	def __writeDataToFile(self, data:pd.DataFrame, settings:dict)-> bool:
		"""
		Write the dataFrame to the requested file

		data:pd.DataFrame = Data to write to the table
		settings:dict = Settings with the path of the file

		->bool = Will return True if successful
		"""

		_fileWritten = False
		_fileType = self.reportFilePath.split(".")[-1]
		try:
			if (_fileType == "xlsx") or (_fileType == "xls"):

				with pd.ExcelWriter(path=(self.reportFilePath), mode="a", if_sheet_exists="overlay") as writer:
					data.to_excel(writer, sheet_name= settings["sheet"], index=False)

				#Create an csv file from the calculated ExcelFile
				self.__createCalculatedCsv(excelFilePath=self.reportFilePath, excelSheet=settings["sheet"], csvSeparator=settings["separator"])

				# Write the file was successful
				_fileWritten = True

			elif (_fileType == "csv"):

				_mode = ""
				if settings["fromTemplate"]:
					_mode = "a"
				else:
					_mode = "w"
				
				data.to_csv(path_or_buf= (self.reportFilePath), mode=_mode, index=False, header=True, sep=settings["separator"])
				
				# Write the file was successful
				_fileWritten = True
			

		except:
			self.logger.exception("Could not write Data to File: " + self.reportFilePath)

		return _fileWritten

	def __getAggregatedDataList(self, eliona:ElionaApiHandler, assetId:int, attribute:str, startDateTime:datetime, 
								endDateTime:datetime, raster:str, mode:str, timeStampKey:str, valueKey:str, assetGai:str="",
								fillNone="") -> tuple[dict|None, pd.DataFrame|None, bool]:
		"""
		Get an attribute value from the given time span with tick
		Will return an dictionary with time stamp as key

		Params
		------
		eliona:ElionaApiHandler	= eliona API Handler instance
		assetId:int = Asset ID to get the data from
		attribute:str = Attribute from the Asset to read the data from
		startDateTime:datetime = Start point from which we create an dictionary entry every time tick  
		endDateTime:datetime = End point for the dictionary
		tick:timedelta = pipeline raster to search for.

		Return
		------
		-> (dict:{datetime, dataValue}|None, bool: True= keys are valid // False = at least one key timestamp is missing)	
		"""

		#Create the return value as a dictionary 
		_dataSet = {}
		_validKeys = True
		_dataFrame = pd.DataFrame(columns=(timeStampKey, valueKey))

		try:

			_assetId = 0
			_utcOffset = int(startDateTime.utcoffset().total_seconds()/3600)
			_startDate = (startDateTime-timedelta(hours=_utcOffset + 1)).isoformat()
			_endDate = (endDateTime+timedelta(hours=_utcOffset + 1)).isoformat()

			if assetGai != "":

				self.logger.debug(f"get data from: assetGai: {assetGai} // attribute: {attribute} // raster: {raster} // start date: {_startDate} // end date: {_endDate}")

				_retVal, part = eliona.get_data_aggregated(	asset_gai=assetGai, 
															from_date=_startDate, 
															to_date=_endDate, 
															data_subtype="input",
															raster=raster,
															attribute=attribute)
				_assetId = eliona.get_asset_id(asset_gai=assetGai)

			elif assetId > 0 :

				self.logger.debug(f"get data from: assetGai: {assetGai} // attribute: {attribute} // raster: {raster} // start date: {_startDate} // end date: {_endDate}")

				_retVal, part = eliona.get_data_aggregated(	asset_id=assetId, 
															from_date=_startDate, 
															to_date=_endDate, 
															data_subtype="input",
															raster=raster,
															attribute=attribute)
				_assetId = assetId

			
			# Dictionary will return True if not empty
			if _retVal:

				#Get the requested data and acquisition mode
				for _data in _retVal:

					#write the info to the LOGGER
					if ((str(_data["asset_id"]) == str(_assetId))
						and (_data["attribute"] == attribute) 
						and (_data["raster"] == raster)
						and mode in _data ):

						_dataSet[_data["timestamp"]] = _data[mode]
						_dataFrame = pd.concat([_dataFrame, pd.DataFrame([[_data["timestamp"].replace(tzinfo=None), _data[mode]]], columns=(timeStampKey, valueKey))] )

						self.logger.debug(	"Timestamp" + str(_data["timestamp"]) + " // AssetId:  " + 
								str(_data["asset_id"]) + " // Attribute: " + str(_data["attribute"]) + 
								" // Raster: " + str(_data["raster"]) + " // Value: " +str(_data[mode]))

				#Validate the Data
				_checkActive = False
				_currentTimeSpan = startDateTime
				_timeDelta = timedelta()
				_missedTimeStamps =  "Asset ID: " + str(_assetId) +  " // Attribute: " + attribute + " // Missed time stamps: "

		
				if raster.find("DAY") != -1:

					_checkActive = False

				elif raster.find("MONTH") != -1:

					_checkActive = False

				elif raster.find("YEAR") != -1:

					_checkActive = False

				elif raster.startswith("M"):

					_timeDelta = timedelta(minutes=int(raster.replace("M", "")))
					_checkActive = True

				elif raster.startswith("H"):

					_timeDelta = timedelta(hours=int(raster.replace("H", "")))
					_checkActive = True

				elif raster.startswith("S"):

					_timeDelta = timedelta(seconds=int(raster.replace("S", "")))
					_checkActive = True


				#Check the TimeStamp
				while _checkActive:

					#Check if we are in range
					if _currentTimeSpan > endDateTime:
						_checkActive = False
					
					#We are in range check if key exists
					elif not _currentTimeSpan in _dataSet:
						#Save the timestamps
						_missedTimeStamps = _missedTimeStamps + "\n	-" + str(_currentTimeSpan) 
						_validKeys = False

					#Cunt up the current time for the next round
					_currentTimeSpan = _currentTimeSpan + _timeDelta

				if not _validKeys:
					self.logger.error(f"Missed Timestamp: {_missedTimeStamps}")

			# Reset valid keys if empty data was received
			else:
				_validKeys = False

		except Exception as err:
			self.logger.exception("Exception getting aggregated data\n" + str(err))
		
		#Return the values
		return (_dataSet, _dataFrame, _validKeys)

	def __readTableTemplate(self, settings:dict) -> pd.DataFrame | None:
		"""
		Read the template Data 

		settings:dict {	"path": "path of the template file",
						"sheet": "Sheet name if excel file is used",
						"type": "file type: csv, xls, xlsx"}
		returns an pandas dataFrame
		"""

		_template = None

		try:
	
			with open(settings["templateFile"], 'r') as tempfile: # OSError if file exists or is invalid

				#Read the file
				if settings["templateFile"].endswith(".csv"):

					_template = pd.read_csv(settings["templateFile"], sep=settings["separator"])

				elif settings["templateFile"].endswith(".xlsx") or settings["templateFile"].endswith(".xls"):

					wb = load_workbook(filename = settings["templateFile"])
					sheet_ranges = wb[settings["sheet"]]
					_template = pd.DataFrame(sheet_ranges.values)

					#set column names equal to values in row index position 0
					_template.columns = _template.iloc[0]

					#remove first row from DataFrame
					_template = _template[1:]
					_template.reset_index(drop=True, inplace=True)

					#_template = pd.read_excel(io=settings["templateFile"], sheet_name=settings["sheet"])

		except OSError:
			self.logger.exception("Template file could not be opened: " + settings["templateFile"])


		#Return the _template
		return _template

	def __createCalculatedCsv(self, excelFilePath:str, excelSheet:str, csvSeparator:str) -> bool:
		"""
		Create a csv file from a excel with possible calculations.
		Will create a template excel file and 

		Params
		----
		excelFilePath:str	= Filepath of the ExelFile


		Return
		----
		Will Return True if everything was created

		"""


		_fileWritten = False
		_fileType = excelFilePath.split(".")[-1]


		try:
			#The variable spreadsheet provides the full path with filename to the excel spreadsheet with unevaluated formulae		
			_fpath = os.path.basename(excelFilePath) 
			_dirname = os.path.dirname(excelFilePath) + "/calculated" 

			_excelModel = formulas.ExcelModel().loads(excelFilePath).finish()
			_excelModel.calculate()
			_excelModel.write( dirpath=(_dirname))

			#Use openpyxl to open the updated excel spreadsheet now
			_wb = load_workbook(filename = _dirname + "/" + _fpath.upper(), data_only = True)
			_sheetRanges = _wb[excelSheet.upper()]
			_calculatedDataFrame = pd.DataFrame(_sheetRanges.values)

			#set column names equal to values in row index position 0
			_calculatedDataFrame.columns = _calculatedDataFrame.iloc[0]

			#remove first row from DataFrame
			_calculatedDataFrame = _calculatedDataFrame[1:]
			_calculatedDataFrame.reset_index(drop=True, inplace=True)

			# Create the csv file 
			_csvReportPath = excelFilePath.replace(_fileType, "csv")
			_calculatedDataFrame.to_csv(path_or_buf=_csvReportPath, mode="w", index=False, header=True, sep=csvSeparator)

			# Write the file was successful
			_fileWritten = True

			# Clean up the temp file
			os.remove(_dirname + "/" + _fpath.upper())


		except:
			self.logger.exception("Could not create csv file from Excel File: " + self.reportFilePath)

		return _fileWritten

	def getLastReceivedValue(self, eliona:ElionaApiHandler, assetGai:str, assetId:int, attribute:str, startDateTime:datetime)->str:
		"""
		Will try to get the last received Value for the given asset and attribute.
		=> try to get the timestamp of this year. If not available try to get the timestamp from last year.
		If No value is available NO-Value will be returned

		Params
		------
		eliona:ElionaApiHandler
		assetGai:str
		assetId:str
		config:dict
		configRaw:dict
		startDateTime:datetime

		"""

		_value = ""
		_loopCounter = 12
		_startTimestamp = startDateTime
		
		while True:

			_endTimestamp = startDateTime
			if startDateTime.month > 1:
				# We will set the day to the first of the Mont. 
				# Otherwise maybe the month will not have the day of the next month
				_startTimestamp = _startTimestamp.replace(month=_startTimestamp.month-1, day=1)
			else:
				_startTimestamp = _startTimestamp.replace(month=12, day=1)

			if assetId == 0:
				_assetId = eliona.get_asset_id(asset_gai=assetGai)
			else:
				_assetId = assetId

			_data, _errorMsg = eliona.get_data_trends(  asset_id=_assetId,
														from_date=_startTimestamp.isoformat(),
														to_date=_endTimestamp.isoformat(),
														data_subtype="input")


			if _errorMsg == "":
				_value = "NO-VALUE"
				if len(_data)>0:
					if _data[-1].get("asset_id", "") == _assetId:
						_lastData = _data[-1].get("data", {})
						_value = _lastData.get(attribute, None)
						break
			else:
				_value = "NO-VALUE"
				_startDateStr = _startTimestamp.isoformat()
				_endDateStr = _endTimestamp.isoformat()
				self.logger.error(f"Tried to read the trend data from: {_startDateStr} to: {_endDateStr} with AssetGai {assetGai} Attribute: {attribute} Error Msg: {_errorMsg}")
				break

			if _loopCounter <= 0:
				break
			else:
				_loopCounter = _loopCounter - 1

		return _value

	def __findJson(self, text:str):
		"""
		Find JSON objects in text, and yield the decoded JSON data

		Does not attempt to look for JSON arrays, text, or other JSON types outside
		of a parent JSON object.

		Params
		----


		"""
		decoder=JSONDecoder()
		pos = 0
		while True:
			match = text.find('{', pos)
			if match == -1:
				break
			try:
				result, index = decoder.raw_decode(text[match:])
				resultRaw = text[match:(match + index)]
				yield result, resultRaw
				pos = match + index
			except ValueError:
				pos = match + 1

	def __isFloat(self, num:str) -> bool:
		"""
		Check if a string is a float or not
		
		Params:
		------
		num		:number as astring to check

		Return:
		------
		Will Return True if string is a valid float. False if not

		"""

		try:
			float(num)
			return True
		except ValueError:
			return False